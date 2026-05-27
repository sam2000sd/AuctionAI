
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import os
import sys
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# Streamlit Cloud runs this file from inside /app, so add the repo root
# to Python path before importing app.core/app.scrapers modules.
ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.config import SCRAPED_DIR, EXPORT_DIR
from app.core.normalize import normalize_files
from app.core.formulas import calc_bid
from app.core.utils import money, pct, this_or_next_week, city_from_address, address_key, zillow_link, redfin_link, redfin_search_link, MD_COUNTIES
from app.scrapers.sources import scrape_many, clear_cache
from app.storage.local import load_bids, save_bids, merge_bids, load_hidden, hide_address, clear_hidden, load_blocked_cities, save_blocked_cities, load_layout_defaults, save_layout_defaults

st.set_page_config(page_title="Auction Intelligence", page_icon="🏛️", layout="wide")

st.markdown("""
<style>
.block-container {max-width: 1900px; padding-top: 1rem;}
.date-header {border-top:5px solid black; background:#f3f4f6; padding:10px 14px; margin-top:24px; font-weight:800;}
.small {color:#6b7280; font-size:.85rem;}
div.stButton > button {white-space: nowrap; min-width: 78px;}
/* make text inputs used for numeric entry match the compact original row shape */
div[data-testid="stTextInput"] input {height: 38px;}
</style>
""", unsafe_allow_html=True)

SCRAPED_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
layout_defaults = load_layout_defaults()

def default_value(key, fallback):
    return layout_defaults.get(key, fallback)


def start_shutdown_listener():
    """Close local Python/Streamlit after the browser tab is really gone."""
    if st.session_state.get("shutdown_listener_started"):
        return
    st.session_state.shutdown_listener_started = True
    state = {"last_ping": datetime.now().timestamp(), "closed_at": None}

    class ShutdownHandler(BaseHTTPRequestHandler):
        def _headers(self):
            self.send_response(204)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
            self.end_headers()

        def do_OPTIONS(self):
            self._headers()

        def do_POST(self):
            path = (self.path or "").lower()
            now = datetime.now().timestamp()
            if "ping" in path:
                state["last_ping"] = now
                state["closed_at"] = None
            elif "closed" in path or "shutdown" in path:
                state["closed_at"] = now
            self._headers()

        def log_message(self, *args):
            return

    def monitor():
        import time
        while True:
            time.sleep(1.0)
            now = datetime.now().timestamp()
            closed_at = state.get("closed_at")
            if closed_at and now - max(closed_at, state.get("last_ping", 0)) > 6:
                os._exit(0)

    def run_server():
        try:
            threading.Thread(target=monitor, daemon=True).start()
            HTTPServer(("127.0.0.1", 8765), ShutdownHandler).serve_forever()
        except OSError:
            pass

    threading.Thread(target=run_server, daemon=True).start()

start_shutdown_listener()
components.html("""
<script>
(function () {
  function ping() {
    try { navigator.sendBeacon('http://127.0.0.1:8765/ping', '1'); } catch(e) {}
  }
  ping();
  setInterval(ping, 2000);
  window.addEventListener('beforeunload', function () {
    try { navigator.sendBeacon('http://127.0.0.1:8765/closed', '1'); } catch(e) {}
  });
})();
</script>
""", height=0)


def paths():
    return sorted([p for p in SCRAPED_DIR.glob("*.csv") if p.stat().st_size > 0])

@st.cache_data(show_spinner=False)
def load_data(path_strings, mtimes):
    return normalize_files([Path(p) for p in path_strings])

def clean_external_url(raw, auctioneer=""):
    """Return a real outside URL only. Prevents Streamlit from opening itself for bad/blank links."""
    url = str(raw or "").strip()
    if not url or url.lower() in {"nan", "none", "#"} or url.startswith("javascript:"):
        return ""
    if url.startswith("//"):
        url = "https:" + url
    if url.startswith("/"):
        base_by_auctioneer = {
            "TW": "https://www.tidewaterauctions.com",
            "AC": "https://realestate.alexcooper.com",
            "HW": "https://www.hwestauctions.com",
            "MWC": "https://apps.mwc-law.com",
        }
        base = base_by_auctioneer.get(str(auctioneer or "").upper(), "")
        url = base + url if base else ""
    low = url.lower()
    if not (low.startswith("http://") or low.startswith("https://")):
        return ""
    # If a bad relative/empty link somehow becomes the Streamlit app URL, suppress it.
    if "streamlit.app" in low or "localhost:8501" in low or "127.0.0.1:8501" in low:
        return ""
    return url

def export_view_df(df):
    """Flat investor-friendly export. Editable numbers stay in thousands."""
    out = pd.DataFrame()
    src = df.copy()
    dt = pd.to_datetime(src.get("Sale Date & Time"), errors="coerce")
    try:
        out["Date/Time"] = dt.dt.strftime("%-m/%-d/%Y %-I:%M %p").fillna("")
    except Exception:
        out["Date/Time"] = dt.dt.strftime("%m/%d/%Y %I:%M %p").fillna("")
    out["Address"] = src.get("Address", "")
    out["Auct"] = src.get("Auctioneer", "")
    out["Note"] = src.get("My Note", "")
    out["Deposit"] = src.get("Deposit", "")
    out["County"] = src.get("County", "")
    out["Occupied"] = src.get("Occupied", False)
    out["Look"] = src.get("Look", "")
    out["Comp"] = pd.to_numeric(src.get("Comp", 0), errors="coerce").fillna(0)
    out["Repair"] = pd.to_numeric(src.get("Rehab", 0), errors="coerce").fillna(0)
    out["Profit"] = pd.to_numeric(src.get("Profit", 0), errors="coerce").fillna(0)
    out["Max"] = ""
    out["%"] = ""
    out["MaxS"] = ""
    out["Ad"] = [clean_external_url(a, au) for a, au in zip(src.get("Ad Link", ""), src.get("Auctioneer", ""))]
    return out

def excel_bytes(df, sale_net=0.96, close1=0.06, close2=0.05):
    """Create a phone-friendly Excel export with live formulas.

    Inputs are in thousands, matching the auction workflow:
      Comp 700, Repair 100, Profit 100 -> Max = ((700*96%)-100-100)/(1+6%).
    User can edit Comp/Repair/Profit on the phone and Max/%/MaxS recalculate.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    bio = BytesIO()
    export_df = export_view_df(df)
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Auctions")
        wb = writer.book
        ws = wb["Auctions"]
        ws.freeze_panes = "A2"

        # Header styling similar to Sam's auction spreadsheet.
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Live formulas. Excel will recalculate when opened/edited.
        # I=Comp, J=Repair, K=Profit, L=Max, M=%, N=MaxS.
        for r in range(2, ws.max_row + 1):
            ws[f"L{r}"] = f'=IF(I{r}>0,((I{r}*{sale_net})-J{r}-K{r})/(1+{close1}),"")'
            ws[f"M{r}"] = f'=IF(I{r}>0,L{r}/I{r},"")'
            ws[f"N{r}"] = f'=IF(I{r}>0,((I{r}*{sale_net})-J{r}-K{r})/(1+{close2}),"")'

            # Clickable Ad link without showing a long URL.
            ad_url = str(ws[f"O{r}"].value or "").strip()
            if ad_url:
                ws[f"O{r}"].hyperlink = ad_url
                ws[f"O{r}"].value = "Ad"
                ws[f"O{r}"].style = "Hyperlink"

        # Formatting.
        widths = {
            "A": 18, "B": 44, "C": 8, "D": 26, "E": 13, "F": 22, "G": 10, "H": 10,
            "I": 10, "J": 10, "K": 10, "L": 12, "M": 8, "N": 12, "O": 10,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in ["I", "J", "K", "L", "N"]:
                ws[f"{col}{row[0].row}"].number_format = '$#,##0'
            ws[f"M{row[0].row}"].number_format = '0%'

        # Make it sortable/filterable on phone/desktop.
        if ws.max_row >= 2:
            end_col = get_column_letter(ws.max_column)
            tab = Table(displayName="AuctionExport", ref=f"A1:{end_col}{ws.max_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)

        # Hidden settings sheet so the assumptions are preserved inside the file.
        settings = wb.create_sheet("Formula Settings")
        settings["A1"] = "Sale net"
        settings["B1"] = sale_net
        settings["A2"] = "Max closing cost"
        settings["B2"] = close1
        settings["A3"] = "MaxS closing cost"
        settings["B3"] = close2
        settings["A4"] = "Formula note"
        settings["B4"] = "Comp, Repair, Profit, Max, and MaxS are in thousands."
        settings.sheet_state = "hidden"

        # Force recalculation in Excel/mobile apps.
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass
    return bio.getvalue()

def safe_key(prefix, aid):
    return prefix + "_" + str(aid).replace("-", "_")

def remember_filter_state():
    keys = [
        "source_filter_sidebar", "hide_hidden_toggle", "hide_blocked_toggle",
        "auctioneer_grid_filter", "county_grid_filter", "search_grid_filter", "date_view_filter",
        "addr_w", "county_w", "note_w", "show_links_toggle", "show_ai_toggle"
    ]
    st.session_state["_restore_filters"] = {k: st.session_state.get(k) for k in keys if k in st.session_state}

def restore_filter_state():
    saved = st.session_state.pop("_restore_filters", None)
    if not saved:
        return
    for k, v in saved.items():
        st.session_state[k] = v

restore_filter_state()

def load_row_state(df, bids):
    if "loaded_ids" not in st.session_state:
        st.session_state.loaded_ids = set()
    latest = {}
    if not bids.empty and "Auction ID" in bids:
        latest = bids.sort_values("Saved At").drop_duplicates("Auction ID", keep="last").set_index("Auction ID").to_dict("index")
    for _, r in df.iterrows():
        aid = str(r["Auction ID"])
        if aid in st.session_state.loaded_ids:
            continue
        b = latest.get(aid, {})
        st.session_state[safe_key("occ", aid)] = bool(b.get("Occupied", r.get("Occupied", False)))
        st.session_state[safe_key("look", aid)] = str(b.get("Look", r.get("Look", "")) or "")
        def _blank_or_int(v):
            try:
                if v is None or str(v).strip().lower() in {"", "nan", "none"}:
                    return ""
                f = float(str(v).replace(",", ""))
                return "" if f == 0 else str(int(f))
            except Exception:
                return ""
        st.session_state[safe_key("comp", aid)] = _blank_or_int(b.get("Comp", r.get("Comp", "")))
        st.session_state[safe_key("rehab", aid)] = _blank_or_int(b.get("Rehab", r.get("Rehab", "")))
        st.session_state[safe_key("profit", aid)] = _blank_or_int(b.get("Profit", r.get("Profit", "")))
        note = b.get("My Note", r.get("My Note", "")) or ""
        if str(note).lower() in {"nan", "none"}:
            note = ""
        st.session_state[safe_key("note", aid)] = str(note)
        st.session_state.loaded_ids.add(aid)

def build_save_df(df, multiplier, sale_net, close1, close2):
    rows = []
    for _, r in df.iterrows():
        aid = str(r["Auction ID"])
        def _num(v):
            try:
                txt = str(v or "").replace(",", "").strip()
                return int(float(txt)) if txt else 0
            except Exception:
                return 0
        comp = _num(st.session_state.get(safe_key("comp", aid), ""))
        rehab = _num(st.session_state.get(safe_key("rehab", aid), ""))
        profit = _num(st.session_state.get(safe_key("profit", aid), ""))
        maxb, bidpct, maxs = calc_bid(comp, rehab, profit, sale_net, close1, close2, multiplier)
        row = r.to_dict()
        row.update({
            "Saved At": datetime.now().isoformat(timespec="seconds"),
            "Occupied": st.session_state.get(safe_key("occ", aid), False),
            "Look": st.session_state.get(safe_key("look", aid), ""),
            "Comp": comp, "Rehab": rehab, "Profit": profit,
            "Max": maxb, "%": bidpct, "MaxS": maxs,
            "My Note": st.session_state.get(safe_key("note", aid), ""),
        })
        rows.append(row)
    return pd.DataFrame(rows)

st.title("Auction Intelligence")
st.caption("Clean rebuild. Local Streamlit dashboard for Maryland/DC foreclosure auctions.")

with st.sidebar:
    st.header("Auction Sources")
    selected = st.multiselect("Sources", ["AC", "TW", "HW", "MWC"], default=default_value("source_filter_sidebar", ["AC", "TW", "HW", "MWC"]), key="source_filter_sidebar")

    c1, c2 = st.columns(2)
    if c1.button("Scrape AC"):
        st.session_state.last_scrape = scrape_many(["AC"], clear_old=False); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()
    if c2.button("Scrape TW"):
        st.session_state.last_scrape = scrape_many(["TW"], clear_old=False); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()
    c3, c4 = st.columns(2)
    if c3.button("Scrape HW"):
        st.session_state.last_scrape = scrape_many(["HW"], clear_old=False); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()
    if c4.button("Scrape MWC"):
        st.session_state.last_scrape = scrape_many(["MWC"], clear_old=False); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()

    if st.button("Full Refresh Selected", type="primary", use_container_width=True):
        st.session_state.last_scrape = scrape_many(selected, clear_old=True); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()
    if st.button("Clear Scraped Cache", use_container_width=True):
        clear_cache(); st.cache_data.clear(); st.session_state.pop("loaded_ids", None); st.rerun()

    if "last_scrape" in st.session_state:
        st.write("Last scrape:")
        for s, res in st.session_state.last_scrape.items():
            if res["ok"]:
                st.success(f"{s}: {res['rows']} rows")
            else:
                st.error(f"{s}: {res['error']}")

    st.write(f"Cache files: **{len(paths())}**")

    st.divider()
    st.header("Focus")
    hide_hidden = st.toggle("Hide hidden properties", default_value("hide_hidden_toggle", True), key="hide_hidden_toggle")
    hide_blocked = st.toggle("Hide blocked cities", default_value("hide_blocked_toggle", True), key="hide_blocked_toggle")
    blocked = st.text_area("Blocked cities, one per line", value="\n".join(sorted(load_blocked_cities())), height=90, key="blocked_cities_text")
    if st.button("Save Blocked Cities", use_container_width=True):
        remember_filter_state()
        save_blocked_cities(set(st.session_state.get("blocked_cities_text", "").splitlines()))
        st.success("Blocked cities saved. Current filters kept.")
        st.rerun()
    if st.button("Clear Hidden Properties", use_container_width=True):
        clear_hidden(); st.rerun()

    st.divider()
    st.header("Formula")
    thousands = st.toggle("Inputs are in thousands", True)
    multiplier = 1000 if thousands else 1
    sale_net = st.number_input("Sale net (%)", 0.0, 100.0, 96.0, .5) / 100
    close1 = st.number_input("Max closing cost (%)", 0.0, 25.0, 6.0, .25) / 100
    close2 = st.number_input("MaxS closing cost (%)", 0.0, 25.0, 5.0, .25) / 100

    st.divider()
    st.header("Optional")
    show_ai = st.toggle("Show Sam AI columns", default_value("show_ai_toggle", False), key="show_ai_toggle")
    show_links = st.toggle("Show Zillow/Redfin links", default_value("show_links_toggle", True), key="show_links_toggle")

    st.divider()
    st.header("Column Widths")
    st.caption("Adjust grid column widths here. Streamlit does not support true drag-resize for this custom editable row layout.")
    addr_w = st.slider("Address width", 1.5, 5.0, float(default_value("addr_w", 2.5)), 0.25, key="addr_w")
    county_w = st.slider("County width", 0.7, 2.5, float(default_value("county_w", 1.0)), 0.1, key="county_w")
    note_w = st.slider("Note width", 0.8, 3.0, float(default_value("note_w", 1.2)), 0.1, key="note_w")

    if st.button("Save Default Layout", use_container_width=True):
        save_layout_defaults({
            "source_filter_sidebar": st.session_state.get("source_filter_sidebar", ["AC", "TW", "HW", "MWC"]),
            "auctioneer_grid_filter": st.session_state.get("auctioneer_grid_filter", []),
            "county_grid_filter": st.session_state.get("county_grid_filter", []),
            "date_view_filter": st.session_state.get("date_view_filter", "Current auction week"),
            "hide_hidden_toggle": st.session_state.get("hide_hidden_toggle", True),
            "hide_blocked_toggle": st.session_state.get("hide_blocked_toggle", True),
            "show_links_toggle": st.session_state.get("show_links_toggle", True),
            "show_ai_toggle": st.session_state.get("show_ai_toggle", False),
            "addr_w": st.session_state.get("addr_w", 2.5),
            "county_w": st.session_state.get("county_w", 1.0),
            "note_w": st.session_state.get("note_w", 1.2),
        })
        st.success("Default layout saved.")


p = paths()
# IMPORTANT: do not auto-run a full scrape on app launch.
# The launcher is portable and may be opened on a new computer; forcing AC/TW/HW/MWC
# scraping immediately can make startup look frozen for many minutes.
# Load cached CSVs instantly, then let the user refresh selected sources manually.
raw = load_data([str(x) for x in p], [x.stat().st_mtime for x in p]) if p else pd.DataFrame()
bids = load_bids()
df = merge_bids(raw, bids) if not raw.empty else raw

if not df.empty:
    df["County"] = df["County"].apply(lambda x: x if x in MD_COUNTIES else "Unknown County")

if df.empty:
    st.warning("No cached auction data found. Select sources in the sidebar and click a scrape/refresh button. First setup on a new computer installs dependencies once, but scraping is manual so startup stays fast.")
    st.stop()

# focus filters
if hide_hidden:
    hidden = load_hidden()
    if hidden:
        df = df[~df["Address"].apply(lambda x: address_key(x) in hidden)]
if hide_blocked:
    blocked_set = load_blocked_cities()
    if blocked_set:
        df = df[~df["Address"].apply(lambda x: city_from_address(x) in blocked_set)]

st.subheader("Filters")
for _lk, _lv in {
    "auctioneer_grid_filter": default_value("auctioneer_grid_filter", []),
    "county_grid_filter": default_value("county_grid_filter", []),
    "date_view_filter": default_value("date_view_filter", "Current auction week"),
}.items():
    if _lk not in st.session_state:
        st.session_state[_lk] = _lv
f1, f2, f3, f4 = st.columns([1, 1.5, 2, 1.3])
auctioneer_filter = f1.multiselect("Auctioneer", sorted(df["Auctioneer"].dropna().unique()), key="auctioneer_grid_filter")
county_values = set(df["County"].dropna().astype(str))
county_options = [c for c in MD_COUNTIES if c in county_values]
county_filter = f2.multiselect("County", county_options, key="county_grid_filter")
search = f3.text_input("Search", key="search_grid_filter")
date_view = f4.radio("Date view", ["Current auction week", "All future", "All dates"], horizontal=False, key="date_view_filter")

filtered = df.copy()
if auctioneer_filter:
    filtered = filtered[filtered["Auctioneer"].isin(auctioneer_filter)]
if county_filter:
    filtered = filtered[filtered["County"].isin(county_filter)]
if search.strip():
    q = search.lower().strip()
    filtered = filtered[(filtered["Address"] + " " + filtered["County"] + " " + filtered["Deposit"]).str.lower().str.contains(q, na=False)]

dates = pd.to_datetime(filtered["Sale Date & Time"], errors="coerce").dt.date
if date_view == "Current auction week":
    a, b = this_or_next_week()
    filtered = filtered[(dates >= a) & (dates <= b)]
elif date_view == "All future":
    from datetime import date
    filtered = filtered[dates >= date.today()]

filtered = filtered.sort_values(["Sale Date & Time", "County", "Address"])
if filtered.empty:
    st.info("No auctions match filters.")
    st.stop()

load_row_state(filtered, bids)
visible_save = build_save_df(filtered, multiplier, sale_net, close1, close2)
if not visible_save.empty:
    existing = load_bids()
    combined = pd.concat([existing, visible_save], ignore_index=True)
    combined = combined.sort_values("Saved At").drop_duplicates("Auction ID", keep="last")
    save_bids(combined)

top1, top2, top3 = st.columns([1,1,4])
if top1.button("Save Bids", type="primary", use_container_width=True):
    save_bids(pd.concat([load_bids(), visible_save], ignore_index=True).sort_values("Saved At").drop_duplicates("Auction ID", keep="last"))
    st.success("Saved")
top2.download_button("Export Excel", data=excel_bytes(visible_save, sale_net, close1, close2), file_name="auction_intelligence.xlsx", use_container_width=True)

st.subheader("Main Auction Grid")
if date_view == "Current auction week":
    a, b = this_or_next_week()
    st.caption(f"Default view: {a.strftime('%m/%d/%Y')} to {b.strftime('%m/%d/%Y')}.")

filtered["_Date"] = pd.to_datetime(filtered["Sale Date & Time"], errors="coerce").dt.date

for d, group in filtered.groupby("_Date", dropna=False):
    st.markdown(f'<div class="date-header">{pd.to_datetime(d).strftime("%A, %B %d, %Y") if pd.notna(d) else "Unknown Date"}</div>', unsafe_allow_html=True)
    widths = [.7,.5,county_w,addr_w,.8,.45,.75,.8,.8,.8,.9,.6,.9,note_w,.45]
    headers = ["Time","Auct","County","Address","Deposit","Occ","Look","Comp","Rehab","Profit","Max","%","MaxS","Note","Ad"]
    if show_ai:
        widths += [.8,.8]
        headers += ["AI ARV","AI Max"]
    if show_links:
        widths += [.8]
        headers += ["Links"]
    widths += [1.35]
    headers += ["Hide"]
    cols = st.columns(widths)
    for c, h in zip(cols, headers):
        c.markdown(f"**{h}**")

    for _, r in group.iterrows():
        aid = str(r["Auction ID"])
        cols = st.columns(widths)
        dt = pd.to_datetime(r["Sale Date & Time"], errors="coerce")
        cols[0].write("" if pd.isna(dt) else dt.strftime("%I:%M %p"))
        cols[1].write(r["Auctioneer"])
        cols[2].write(r["County"])
        cols[3].write(r["Address"])
        cols[4].write(r["Deposit"])
        cols[5].checkbox("Occ", key=safe_key("occ", aid), label_visibility="collapsed")
        cols[6].selectbox("Look", ["", "Y", "N", "YY", "Soso"], key=safe_key("look", aid), label_visibility="collapsed")
        cols[7].text_input("Comp", key=safe_key("comp", aid), label_visibility="collapsed", placeholder="")
        cols[8].text_input("Rehab", key=safe_key("rehab", aid), label_visibility="collapsed", placeholder="")
        cols[9].text_input("Profit", key=safe_key("profit", aid), label_visibility="collapsed", placeholder="")
        def _to_int(v):
            try:
                txt = str(v or "").replace(",", "").strip()
                return int(float(txt)) if txt else 0
            except Exception:
                return 0
        comp = _to_int(st.session_state.get(safe_key("comp", aid), ""))
        rehab = _to_int(st.session_state.get(safe_key("rehab", aid), ""))
        profit = _to_int(st.session_state.get(safe_key("profit", aid), ""))
        maxb, bidpct, maxs = calc_bid(comp, rehab, profit, sale_net, close1, close2, multiplier)
        cols[10].write(money(maxb))
        cols[11].write(pct(bidpct))
        cols[12].write(money(maxs))
        cols[13].text_input("Note", key=safe_key("note", aid), label_visibility="collapsed")
        link = clean_external_url(r.get("Ad Link", ""), r.get("Auctioneer", ""))
        cols[14].markdown(f'<a href="{link}" target="_blank" rel="noopener noreferrer">Ad</a>' if link else "", unsafe_allow_html=True)
        idx = 15
        if show_ai:
            # Starter AI: use current comp as AI ARV, and current calculated max as AI Max.
            cols[idx].write(money(float(comp or 0) * multiplier) if comp else "")
            cols[idx+1].write(money(maxb))
            idx += 2
        if show_links:
            addr = r["Address"]
            cols[idx].markdown(f'<a href="{zillow_link(addr)}" target="_blank" rel="noopener noreferrer">Z</a> / <a href="{redfin_search_link(addr)}" target="_blank" rel="noopener noreferrer">R</a>', unsafe_allow_html=True)
            idx += 1
        if cols[idx].button("Hide", key=safe_key("hide", aid), use_container_width=True):
            hide_address(r["Address"])
            st.rerun()

st.divider()
tab1, tab2 = st.tabs(["Saved Bid Archive", "AI Coach"])
with tab1:
    arch = load_bids()
    if arch.empty:
        st.info("No saved bids.")
    else:
        arch_view = arch.drop(columns=["Zestimate", "Redfin Estimate"], errors="ignore")
        st.dataframe(arch_view.sort_values("Saved At", ascending=False), hide_index=True, use_container_width=True)
with tab2:
    arch = load_bids()
    if arch.empty:
        st.info("Save bids first.")
    else:
        comp = pd.to_numeric(arch["Comp"], errors="coerce").fillna(0)
        maxb = pd.to_numeric(arch["Max"], errors="coerce").fillna(0)
        valid = comp > 0
        st.metric("Bids Learned", len(arch))
        st.metric("Average Bid %", pct((maxb[valid] / (comp[valid] * multiplier)).mean()) if valid.any() else "")
